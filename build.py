#!/usr/bin/env python3
"""
Build script: reads dsa-critics-meps-verified.xlsx
  → data/dsa-critics-meps.json
  → data/dsa-critics-meps-verified.xlsx (copy)
  → index.html
"""

import openpyxl
import json
import os
import shutil
from collections import Counter

# ── 1. Read spreadsheet ───────────────────────────────────────────────────────
wb = openpyxl.load_workbook('dsa-critics-meps-verified.xlsx')
ws = wb.active

meps = []
for row in range(2, ws.max_row + 1):
    name_val = ws.cell(row=row, column=1).value
    if not name_val:
        continue
    mep = {
        'name':          str(name_val).strip(),
        'country':       (str(ws.cell(row=row, column=2).value or '').strip() or None),
        'national_party':(str(ws.cell(row=row, column=3).value or '').strip() or None),
        'ep_group':      (str(ws.cell(row=row, column=4).value or '').strip() or None),
        'key_quote':     (str(ws.cell(row=row, column=5).value or '').strip() or None),
        'notes':         (str(ws.cell(row=row, column=6).value or '').strip() or None),
        'sources': []
    }
    for col in range(7, 13):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            mep['sources'].append({
                'label': str(cell.value).strip(),
                'url':   cell.hyperlink.target if cell.hyperlink else None
            })
    meps.append(mep)

print(f"Read {len(meps)} MEPs from spreadsheet")

# ── 2. Save JSON + copy spreadsheet ──────────────────────────────────────────
os.makedirs('data', exist_ok=True)
with open('data/dsa-critics-meps.json', 'w', encoding='utf-8') as f:
    json.dump(meps, f, ensure_ascii=False, indent=2)
print(f"Saved data/dsa-critics-meps.json")

shutil.copy('dsa-critics-meps-verified.xlsx', 'data/dsa-critics-meps-verified.xlsx')
print("Copied spreadsheet to data/")

# ── 3. Pre-compute stats for template ─────────────────────────────────────────
DATA_JS       = json.dumps(meps, ensure_ascii=False, separators=(',', ':'))
N_MEPS        = len(meps)
N_COUNTRIES   = len(set(m['country'] for m in meps if m['country']))
N_GROUPS      = len(set(m['ep_group'] for m in meps if m['ep_group']))

GROUP_COLORS = {
    'PfE': '#BE123C', 'ECR': '#C2410C', 'ESN': '#7F1D1D',
    'EPP': '#1D4ED8', 'Greens/EFA': '#15803D', 'NI': '#6B7280', 'Renew': '#F59E0B'
}
COUNTRY_NAMES = {
    'DE': 'Germany', 'FR': 'France', 'HR': 'Croatia', 'HU': 'Hungary',
    'SK': 'Slovakia', 'AT': 'Austria', 'BE': 'Belgium', 'CZ': 'Czech Republic',
    'PL': 'Poland', 'PT': 'Portugal', 'SI': 'Slovenia', 'BG': 'Bulgaria',
    'ES': 'Spain', 'NL': 'Netherlands', 'SE': 'Sweden', 'RO': 'Romania'
}

groups_by_count = sorted(
    set(m['ep_group'] for m in meps if m['ep_group']),
    key=lambda g: -sum(1 for m in meps if m['ep_group'] == g)
)
pills_html = '\n'.join(
    f'<button class="group-pill" data-group="{g}" '
    f'style="--pill-color:{GROUP_COLORS.get(g, "#6B7280")}">'
    f'{g}<span class="pill-count">{sum(1 for m in meps if m["ep_group"]==g)}</span></button>'
    for g in groups_by_count
)

countries_ranked = Counter(m['country'] for m in meps if m['country']).most_common()
country_options = '<option value="all">All Countries</option>\n' + '\n'.join(
    f'<option value="{c}">{COUNTRY_NAMES.get(c, c)} ({n})</option>'
    for c, n in countries_ranked
)

# ── 4. HTML template (uses %PLACEHOLDER% so no brace-escaping needed) ─────────
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>MEPs Characterizing the DSA as Censorship</title>
<style>
:root {
  --bg: #F8F9FA;
  --surface: #FFFFFF;
  --text: #212529;
  --text-muted: #6B7280;
  --accent: #0D6EFD;
  --accent-hover: #0B5ED7;
  --secondary: #E07A5F;
  --border: #E9ECEF;
  --shadow: 0 1px 3px rgba(0,0,0,.08), 0 1px 2px rgba(0,0,0,.05);
  --radius: 8px;
  --font: system-ui, -apple-system, "Segoe UI", Roboto, sans-serif;
}
*, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
body { font-family: var(--font); background: var(--bg); color: var(--text); line-height: 1.6; }
a { color: var(--accent); text-decoration: none; }
a:hover { text-decoration: underline; }

.container { max-width: 1200px; margin: 0 auto; padding: 0 20px; }

/* ── HEADER ── */
.site-header { background: var(--surface); border-bottom: 1px solid var(--border); padding: 32px 0 24px; }
.site-header h1 { font-size: clamp(1.3rem, 3vw, 2rem); font-weight: 700; color: #1A1A2E; line-height: 1.25; margin-bottom: 10px; }
.subtitle { font-size: 0.97rem; color: var(--text-muted); max-width: 720px; line-height: 1.65; margin-bottom: 20px; }
.stats-bar { display: flex; gap: 24px; flex-wrap: wrap; }
.stat-chip { display: flex; flex-direction: column; }
.stat-number { font-size: 1.6rem; font-weight: 700; color: var(--accent); line-height: 1; }
.stat-label  { font-size: 0.75rem; text-transform: uppercase; letter-spacing: .05em; color: var(--text-muted); margin-top: 2px; }

/* ── FILTER BAR ── */
.filter-bar { background: var(--surface); border-bottom: 1px solid var(--border); padding: 14px 0; position: sticky; top: 0; z-index: 100; }
.filter-inner { display: flex; flex-wrap: wrap; gap: 10px; align-items: center; }
.filter-label { font-size: 0.78rem; font-weight: 600; text-transform: uppercase; letter-spacing: .06em; color: var(--text-muted); white-space: nowrap; }
.group-pills { display: flex; flex-wrap: wrap; gap: 6px; }
.group-pill {
  padding: 4px 10px; border-radius: 20px; border: 2px solid var(--pill-color);
  background: transparent; color: var(--pill-color); font-size: 0.8rem; font-weight: 600;
  cursor: pointer; transition: all .15s; display: flex; align-items: center; gap: 5px;
}
.group-pill:hover { background: var(--pill-color); color: #fff; }
.group-pill.active { background: var(--pill-color); color: #fff; }
.pill-count { font-size: 0.7rem; opacity: .85; }
.filter-divider { width: 1px; height: 28px; background: var(--border); }
select.filter-select {
  padding: 5px 10px; border: 1px solid var(--border); border-radius: 6px;
  font-size: 0.85rem; font-family: var(--font); background: var(--surface);
  color: var(--text); cursor: pointer;
}
.search-wrap { position: relative; }
.search-wrap svg { position: absolute; left: 8px; top: 50%; transform: translateY(-50%); color: var(--text-muted); pointer-events: none; }
input.filter-search {
  padding: 5px 10px 5px 30px; border: 1px solid var(--border); border-radius: 6px;
  font-size: 0.85rem; font-family: var(--font); width: 180px;
}
input.filter-search:focus { outline: 2px solid var(--accent); border-color: transparent; }
.btn-clear {
  padding: 5px 12px; border: 1px solid var(--border); border-radius: 6px;
  background: transparent; font-size: 0.8rem; cursor: pointer; color: var(--text-muted);
  transition: all .15s;
}
.btn-clear:hover { background: var(--bg); color: var(--text); }

/* ── VIZ ROW ── */
.viz-row { display: grid; grid-template-columns: 1fr 340px; gap: 20px; margin: 24px 0; }
@media (max-width: 900px) { .viz-row { grid-template-columns: 1fr; } }
.viz-panel { background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow); padding: 20px; }
.viz-panel h3 { font-size: 0.85rem; font-weight: 700; text-transform: uppercase; letter-spacing: .06em; color: var(--text-muted); margin-bottom: 14px; }

/* Map */
#map-svg { display: block; width: 100%; }
.country { transition: fill .2s; }
.country:hover { fill: #BFDBFE !important; }
.map-fallback { color: var(--text-muted); font-size: 0.9rem; padding: 60px 0; text-align: center; }

/* Charts */
.charts-panel { display: flex; flex-direction: column; gap: 20px; }
.chart-svg { display: block; width: 100%; overflow: visible; }

/* ── RESULTS BAR ── */
.results-bar { display: flex; align-items: center; justify-content: space-between; margin-bottom: 14px; flex-wrap: wrap; gap: 8px; }
#results-count { font-size: 0.85rem; color: var(--text-muted); }
.sort-label { font-size: 0.82rem; color: var(--text-muted); }

/* ── CARD GRID ── */
.card-grid { display: grid; grid-template-columns: repeat(auto-fill, minmax(320px, 1fr)); gap: 16px; margin-bottom: 40px; }
.mep-card { background: var(--surface); border-radius: var(--radius); box-shadow: var(--shadow); padding: 18px 20px; border: 1px solid var(--border); transition: box-shadow .15s; }
.mep-card:hover { box-shadow: 0 4px 12px rgba(0,0,0,.1); }
.card-top { display: flex; align-items: flex-start; justify-content: space-between; gap: 8px; margin-bottom: 6px; }
.card-name { font-weight: 700; font-size: 1.0rem; color: #1A1A2E; line-height: 1.3; }
.card-badges { display: flex; flex-wrap: wrap; gap: 5px; align-items: center; margin-top: 2px; }
.country-tag { font-size: 0.78rem; color: var(--text-muted); }
.group-badge { font-size: 0.72rem; font-weight: 700; padding: 2px 8px; border-radius: 12px; color: #fff; white-space: nowrap; }
.card-party { font-size: 0.83rem; color: var(--text-muted); margin-bottom: 8px; }
.card-quote { font-style: italic; font-size: 0.88rem; color: #374151; border-left: 3px solid var(--secondary); padding-left: 10px; margin: 10px 0; line-height: 1.5; }
.card-notes { font-size: 0.82rem; color: var(--text-muted); margin-bottom: 8px; }
.card-sources { margin-top: 12px; padding-top: 10px; border-top: 1px solid var(--border); }
.sources-label { font-size: 0.72rem; font-weight: 700; text-transform: uppercase; letter-spacing: .06em; color: var(--text-muted); margin-bottom: 5px; }
.card-sources a, .card-sources .src-nohref {
  display: inline-block; font-size: 0.78rem; margin: 2px 5px 2px 0;
  padding: 2px 7px; background: #EFF6FF; border-radius: 4px;
  color: var(--accent); border: 1px solid #DBEAFE; white-space: nowrap;
  max-width: 100%; overflow: hidden; text-overflow: ellipsis;
}
.card-sources a:hover { background: #DBEAFE; text-decoration: none; }
.card-sources .src-nohref { color: var(--text-muted); background: #F3F4F6; border-color: var(--border); }

/* No results */
.no-results { text-align: center; padding: 60px 20px; color: var(--text-muted); grid-column: 1/-1; }
.no-results strong { display: block; font-size: 1.1rem; margin-bottom: 6px; color: var(--text); }

/* ── COUNTRY CHART ── */
.bar-label { font-size: 11px; fill: #374151; font-family: var(--font); }
.bar-count { font-size: 11px; fill: #374151; font-family: var(--font); }

/* ── METHODOLOGY ── */
.methodology { background: var(--surface); border-top: 1px solid var(--border); padding: 40px 0; margin-top: 20px; }
.methodology h2 { font-size: 1.0rem; font-weight: 700; margin-bottom: 14px; color: #1A1A2E; }
.methodology p, .methodology li { font-size: 0.88rem; color: var(--text-muted); line-height: 1.7; margin-bottom: 8px; }
.methodology ul { padding-left: 20px; }
.methodology ul li { margin-bottom: 4px; }
.methodology a { color: var(--accent); }

/* ── RESPONSIVE ── */
@media (max-width: 640px) {
  .filter-inner { gap: 8px; }
  input.filter-search { width: 140px; }
  .card-grid { grid-template-columns: 1fr; }
  .stats-bar { gap: 16px; }
}
</style>
</head>
<body>

<!-- HEADER -->
<header class="site-header">
<div class="container">
  <h1>MEPs Characterizing the DSA as Censorship</h1>
  <p class="subtitle">European Parliament members who have signed parliamentary questions, participated in conferences, or delivered plenary statements framing the Digital Services Act as a vehicle for censorship or an instrument for suppressing free speech. Every source document was fetched and read in full to verify its subject matter.</p>
  <div class="stats-bar">
    <div class="stat-chip"><span class="stat-number" id="stat-meps">%N_MEPS%</span><span class="stat-label">MEPs</span></div>
    <div class="stat-chip"><span class="stat-number" id="stat-countries">%N_COUNTRIES%</span><span class="stat-label">Countries</span></div>
    <div class="stat-chip"><span class="stat-number" id="stat-groups">%N_GROUPS%</span><span class="stat-label">EP Groups</span></div>
  </div>
</div>
</header>

<!-- FILTER BAR -->
<div class="filter-bar">
<div class="container">
<div class="filter-inner">
  <span class="filter-label">Filter by</span>
  <div class="group-pills" id="group-pills">%PILLS_HTML%</div>
  <div class="filter-divider"></div>
  <select class="filter-select" id="filter-country">%COUNTRY_OPTIONS%</select>
  <div class="search-wrap">
    <svg width="14" height="14" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="2"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
    <input type="text" class="filter-search" id="filter-search" placeholder="Search name…" autocomplete="off">
  </div>
  <button class="btn-clear" id="btn-clear">Clear all</button>
</div>
</div>
</div>

<!-- MAIN -->
<main>
<div class="container">

  <!-- VIZ ROW -->
  <div class="viz-row">
    <div class="viz-panel">
      <h3>Geographic Distribution <span style="font-size:.75em;font-weight:400;color:#9CA3AF">(click bubble to filter)</span></h3>
      <svg id="map-svg" height="320"></svg>
    </div>
    <div class="charts-panel">
      <div class="viz-panel">
        <h3>By EP Group</h3>
        <svg id="group-chart-svg" class="chart-svg" height="160"></svg>
      </div>
      <div class="viz-panel">
        <h3>By Country (top 10)</h3>
        <svg id="country-chart-svg" class="chart-svg" height="200"></svg>
      </div>
    </div>
  </div>

  <!-- RESULTS -->
  <div class="results-bar">
    <span id="results-count"></span>
  </div>
  <div class="card-grid" id="mep-grid"></div>

</div>
</main>

<!-- METHODOLOGY -->
<footer class="methodology">
<div class="container">
  <h2>Methodology</h2>
  <p>This dataset identifies European Parliament members who have, in verifiable parliamentary documents or public statements, characterized the EU Digital Services Act (DSA) as a censorship mechanism or instrument for suppressing free speech. The following source types were used:</p>
  <ul>
    <li>Oral questions (O-) filed with the European Commission or Council</li>
    <li>Written questions (E-, P-) filed to the European Commission</li>
    <li>Plenary speeches and debates (verbatim record, January 2025 enforcement debate)</li>
    <li>Public conference participation (ADF International, Brussels Report)</li>
    <li>Collective expert letters and statements citing the DSA</li>
  </ul>
  <p><strong>Excluded documents:</strong></p>
  <ul>
    <li><strong>O-046/2025</strong> — Child protection online. While filed by overlapping signatories, the framing is distinct and was excluded.</li>
    <li><strong>O-042/2025</strong> — "Protecting citizens' right to make cash payments." Not DSA-related; excluded.</li>
    <li><strong>O-043/2025</strong> — "Manipulation of the Erasmus+ programme." Not DSA-related; excluded.</li>
    <li><strong>O-003/2026</strong> — Concerns Reporters Without Borders. Not DSA-related; excluded.</li>
    <li><strong>E-005041/2025</strong> — Telegram/Durov prosecution. Not a DSA-censorship framing; excluded.</li>
  </ul>
  <p>Every source document was fetched from the European Parliament's official document portal and read in full before being attributed to any MEP. The original dataset contained 114 MEPs; after full verification, 55 MEPs were confirmed with at least one qualifying source. 59 MEPs were removed due to reliance on wrongly-attributed source documents.</p>
  <p>Source data: <a href="data/dsa-critics-meps.json">dsa-critics-meps.json</a> &mdash;
     GitHub: <a href="https://github.com/your-username/dsa-critic-map" target="_blank" rel="noopener">your-username/dsa-critic-map</a> &mdash;
     License: <a href="https://creativecommons.org/licenses/by/4.0/" target="_blank" rel="noopener">CC BY 4.0</a></p>
</div>
</footer>

<script src="https://cdnjs.cloudflare.com/ajax/libs/d3/7.9.0/d3.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/topojson-client@3/dist/topojson-client.min.js"></script>
<script>
const DATA = %DATA_JS%;

const GROUP_COLORS = {
  'PfE':'#BE123C','ECR':'#C2410C','ESN':'#7F1D1D',
  'EPP':'#1D4ED8','Greens/EFA':'#15803D','NI':'#6B7280','Renew':'#F59E0B'
};
const COUNTRY_FLAGS = {
  'DE':'🇩🇪','FR':'🇫🇷','HR':'🇭🇷','HU':'🇭🇺','SK':'🇸🇰','AT':'🇦🇹',
  'BE':'🇧🇪','CZ':'🇨🇿','PL':'🇵🇱','PT':'🇵🇹','SI':'🇸🇮','BG':'🇧🇬',
  'ES':'🇪🇸','NL':'🇳🇱','SE':'🇸🇪','RO':'🇷🇴'
};
const COUNTRY_NAMES = {
  'DE':'Germany','FR':'France','HR':'Croatia','HU':'Hungary','SK':'Slovakia',
  'AT':'Austria','BE':'Belgium','CZ':'Czech Republic','PL':'Poland','PT':'Portugal',
  'SI':'Slovenia','BG':'Bulgaria','ES':'Spain','NL':'Netherlands','SE':'Sweden','RO':'Romania'
};
const NUMERIC_TO_ALPHA2 = {
  276:'DE',250:'FR',191:'HR',348:'HU',703:'SK',40:'AT',56:'BE',203:'CZ',
  616:'PL',620:'PT',705:'SI',100:'BG',724:'ES',528:'NL',752:'SE',642:'RO'
};
const COUNTRY_CENTROIDS = {
  'DE':[10.4,51.2],'FR':[2.2,46.2],'HR':[15.9,45.8],'HU':[19.5,47.2],
  'SK':[19.7,48.7],'AT':[14.6,47.5],'BE':[4.5,50.5],'CZ':[15.5,49.8],
  'PL':[19.1,51.9],'PT':[-8.2,39.4],'SI':[15.0,46.1],'BG':[25.5,42.7],
  'ES':[-3.7,40.4],'NL':[5.3,52.1],'SE':[18.6,60.1],'RO':[24.9,45.9]
};
const EUROPEAN_IDS = new Set([
  40,8,20,112,56,70,100,191,196,203,208,233,246,250,276,300,348,352,372,380,
  428,438,440,442,470,498,492,499,528,807,578,616,620,642,643,674,688,703,705,
  724,752,756,792,804,826,336
]);

// ── STATE ──────────────────────────────────────────────────────────────────
const state = { groups: new Set(), country: 'all', search: '' };
let currentFiltered = [...DATA];
const mepCountByCountry = {};
DATA.forEach(m => { if (m.country) mepCountByCountry[m.country] = (mepCountByCountry[m.country]||0)+1; });

// ── HELPERS ────────────────────────────────────────────────────────────────
function esc(s) {
  if (!s) return '';
  return String(s).replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;');
}

function getFiltered() {
  return DATA.filter(m => {
    if (state.groups.size > 0 && !state.groups.has(m.ep_group)) return false;
    if (state.country !== 'all' && m.country !== state.country) return false;
    if (state.search) {
      const q = state.search.toLowerCase();
      if (!(m.name||'').toLowerCase().includes(q) &&
          !(m.national_party||'').toLowerCase().includes(q) &&
          !(m.ep_group||'').toLowerCase().includes(q)) return false;
    }
    return true;
  });
}

// ── CARDS ──────────────────────────────────────────────────────────────────
function renderCards(filtered) {
  const grid = document.getElementById('mep-grid');
  document.getElementById('results-count').textContent =
    `Showing ${filtered.length} of ${DATA.length} MEPs`;

  if (filtered.length === 0) {
    grid.innerHTML = '<div class="no-results"><strong>No MEPs match these filters</strong>Try adjusting your search or clearing filters.</div>';
    return;
  }

  grid.innerHTML = filtered.map(m => {
    const color = GROUP_COLORS[m.ep_group] || '#6B7280';
    const flag  = COUNTRY_FLAGS[m.country] || '';
    const srcs  = m.sources.map(s =>
      s.url
        ? `<a href="${esc(s.url)}" target="_blank" rel="noopener noreferrer" title="${esc(s.label)}">${esc(s.label)}</a>`
        : `<span class="src-nohref">${esc(s.label)}</span>`
    ).join('');

    return `<div class="mep-card">
      <div class="card-top">
        <div>
          <div class="card-name">${esc(m.name)}</div>
          <div class="card-badges">
            <span class="country-tag">${flag} ${esc(m.country||'')}</span>
            ${m.ep_group ? `<span class="group-badge" style="background:${color}">${esc(m.ep_group)}</span>` : ''}
          </div>
        </div>
      </div>
      ${m.national_party ? `<div class="card-party">${esc(m.national_party)}</div>` : ''}
      ${m.key_quote ? `<blockquote class="card-quote">"${esc(m.key_quote)}"</blockquote>` : ''}
      ${m.notes ? `<div class="card-notes">${esc(m.notes)}</div>` : ''}
      ${srcs ? `<div class="card-sources"><div class="sources-label">Sources</div>${srcs}</div>` : ''}
    </div>`;
  }).join('');
}

// ── BAR CHARTS (vanilla SVG) ───────────────────────────────────────────────
function renderBarChart(svgId, entries, colorFn) {
  const svg = document.getElementById(svgId);
  const W = svg.parentElement.clientWidth - 40 || 280;
  const BAR_H = 22, GAP = 5, LABEL_W = 65, PAD = 24;
  const total = entries.length;
  const H = total * (BAR_H + GAP) + PAD;
  svg.setAttribute('width', W);
  svg.setAttribute('height', H);
  svg.setAttribute('viewBox', `0 0 ${W} ${H}`);
  svg.innerHTML = '';

  if (entries.length === 0) return;
  const maxVal = Math.max(...entries.map(([,v]) => v));
  const barMax = W - LABEL_W - 36;

  entries.forEach(([key, val], i) => {
    const y = i * (BAR_H + GAP) + 8;
    const bw = val > 0 ? Math.max(4, (val / maxVal) * barMax) : 0;
    const color = colorFn(key);

    const lbl = document.createElementNS('http://www.w3.org/2000/svg','text');
    lbl.setAttribute('class','bar-label');
    lbl.setAttribute('x', LABEL_W - 5);
    lbl.setAttribute('y', y + BAR_H/2 + 4);
    lbl.setAttribute('text-anchor','end');
    lbl.textContent = key;
    svg.appendChild(lbl);

    const rect = document.createElementNS('http://www.w3.org/2000/svg','rect');
    rect.setAttribute('x', LABEL_W);
    rect.setAttribute('y', y);
    rect.setAttribute('width', bw);
    rect.setAttribute('height', BAR_H);
    rect.setAttribute('fill', color);
    rect.setAttribute('rx', 3);
    svg.appendChild(rect);

    const cnt = document.createElementNS('http://www.w3.org/2000/svg','text');
    cnt.setAttribute('class','bar-count');
    cnt.setAttribute('x', LABEL_W + bw + 5);
    cnt.setAttribute('y', y + BAR_H/2 + 4);
    cnt.textContent = val;
    svg.appendChild(cnt);
  });
}

function renderGroupChart(filtered) {
  const counts = {};
  filtered.forEach(m => { if (m.ep_group) counts[m.ep_group] = (counts[m.ep_group]||0)+1; });
  const entries = Object.entries(counts).sort((a,b)=>b[1]-a[1]);
  renderBarChart('group-chart-svg', entries, g => GROUP_COLORS[g]||'#6B7280');
}

function renderCountryChart(filtered) {
  const counts = {};
  filtered.forEach(m => { if (m.country) counts[m.country] = (counts[m.country]||0)+1; });
  const entries = Object.entries(counts).sort((a,b)=>b[1]-a[1]).slice(0,10)
    .map(([c,v]) => [COUNTRY_NAMES[c]||c, v]);
  const H = entries.length * 27 + 24;
  document.getElementById('country-chart-svg').style.height = H+'px';
  renderBarChart('country-chart-svg', entries, () => '#64748B');
}

// ── D3 MAP ─────────────────────────────────────────────────────────────────
async function renderMap() {
  const mapSvg = document.getElementById('map-svg');
  const W = (mapSvg.parentElement.clientWidth||500) - 40;
  const H = 320;
  mapSvg.setAttribute('width', W);
  mapSvg.setAttribute('viewBox', `0 0 ${W} ${H}`);

  const projection = d3.geoMercator().fitExtent(
    [[10,10],[W-10,H-10]],
    { type:'Feature', geometry:{ type:'Polygon',
      coordinates:[[[-14,33],[35,33],[35,71],[8,71],[-14,65],[-14,33]]] }}
  );
  const pathGen = d3.geoPath().projection(projection);
  const svg = d3.select(mapSvg);

  try {
    const world = await fetch('https://cdn.jsdelivr.net/npm/world-atlas@2/countries-110m.json')
      .then(r => r.json());
    const countries = topojson.feature(world, world.objects.countries);

    svg.selectAll('.country')
      .data(countries.features.filter(f => EUROPEAN_IDS.has(+f.id)))
      .join('path')
      .attr('class','country')
      .attr('d', d => pathGen(d)||'')
      .attr('fill', d => {
        const a2 = NUMERIC_TO_ALPHA2[+d.id];
        return (a2 && mepCountByCountry[a2]) ? '#DBEAFE' : '#E5E7EB';
      })
      .attr('stroke','#fff').attr('stroke-width',0.6);

    const maxCount = Math.max(...Object.values(mepCountByCountry));
    const rScale = d3.scaleSqrt().domain([0,maxCount]).range([0,24]);

    Object.entries(COUNTRY_CENTROIDS).forEach(([code,[lng,lat]]) => {
      const count = mepCountByCountry[code];
      if (!count) return;
      const pt = projection([lng,lat]);
      if (!pt) return;
      const [x,y] = pt;

      const g = svg.append('g').attr('class','bubble-g').style('cursor','pointer')
        .on('click', () => {
          document.getElementById('filter-country').value = code;
          state.country = code;
          applyFilters();
        });

      g.append('title').text(`${COUNTRY_NAMES[code]||code}: ${count} MEPs`);
      g.append('circle')
        .attr('cx',x).attr('cy',y).attr('r',rScale(count))
        .attr('fill','#0D6EFD').attr('fill-opacity',0.78)
        .attr('stroke','#fff').attr('stroke-width',1.5);
      g.append('text')
        .attr('x',x).attr('y',y+4)
        .attr('text-anchor','middle')
        .attr('font-size', Math.max(9, rScale(count)*0.65)+'px')
        .attr('font-weight','700').attr('fill','#fff')
        .attr('pointer-events','none')
        .text(count);
    });
  } catch(e) {
    svg.append('text').attr('x',W/2).attr('y',H/2)
      .attr('text-anchor','middle').attr('fill','#9CA3AF').attr('font-size','13px')
      .text('Map requires internet connection');
  }
}

// ── APPLY FILTERS ──────────────────────────────────────────────────────────
function applyFilters() {
  currentFiltered = getFiltered();
  renderCards(currentFiltered);
  renderGroupChart(currentFiltered);
  renderCountryChart(currentFiltered);
  // Sync pill active state
  document.querySelectorAll('.group-pill').forEach(btn => {
    btn.classList.toggle('active', state.groups.has(btn.dataset.group));
  });
}

function toggleGroup(g) {
  if (state.groups.has(g)) state.groups.delete(g);
  else state.groups.add(g);
  applyFilters();
}

// ── EVENT LISTENERS ────────────────────────────────────────────────────────
document.getElementById('filter-country').addEventListener('change', e => {
  state.country = e.target.value;
  applyFilters();
});
document.getElementById('filter-search').addEventListener('input', e => {
  state.search = e.target.value.trim();
  applyFilters();
});
document.getElementById('btn-clear').addEventListener('click', () => {
  state.groups.clear();
  state.country = 'all';
  state.search = '';
  document.getElementById('filter-country').value = 'all';
  document.getElementById('filter-search').value = '';
  applyFilters();
});
document.querySelectorAll('.group-pill').forEach(btn => {
  btn.addEventListener('click', () => toggleGroup(btn.dataset.group));
});

// ── INIT ───────────────────────────────────────────────────────────────────
document.addEventListener('DOMContentLoaded', () => {
  applyFilters();
  renderMap();
});
</script>
</body>
</html>"""

# ── 5. Perform substitutions ──────────────────────────────────────────────────
html = HTML_TEMPLATE \
    .replace('%DATA_JS%',       DATA_JS) \
    .replace('%N_MEPS%',        str(N_MEPS)) \
    .replace('%N_COUNTRIES%',   str(N_COUNTRIES)) \
    .replace('%N_GROUPS%',      str(N_GROUPS)) \
    .replace('%PILLS_HTML%',    pills_html) \
    .replace('%COUNTRY_OPTIONS%', country_options)

with open('index.html', 'w', encoding='utf-8') as f:
    f.write(html)
print("Generated index.html")
print(f"\nSummary: {N_MEPS} MEPs, {N_COUNTRIES} countries, {N_GROUPS} EP groups")
